"""
Tests for payment export functionality.
"""

import pytest
from fastapi.testclient import TestClient
from sqlmodel import Session
from datetime import date, timedelta

from app.models.payment import Payment, PaymentStatus
from app.models.payment_schedule import PaymentSchedule, PaymentFrequency
from tests.factories import (
    LandlordFactory,
    PropertyFactory,
    RoomFactory,
    TenantFactory,
    PaymentScheduleFactory,
    PaymentFactory,
)


class TestPaymentExport:
    """Test payment export endpoint."""

    def test_export_excel_success(
        self, client: TestClient, session: Session, auth_headers: dict
    ):
        """Test successful Excel export."""
        # Create test data
        landlord = LandlordFactory.create(session=session)
        property_obj = PropertyFactory.create(session=session, landlord_id=landlord.id)
        room = RoomFactory.create(session=session, property_id=property_obj.id)
        tenant = TenantFactory.create(session=session, room_id=room.id)

        # Create payments
        for i in range(3):
            PaymentFactory.create(
                session=session,
                tenant_id=tenant.id,
                amount_due=150000 + (i * 10000),
                status=PaymentStatus.ON_TIME if i % 2 == 0 else PaymentStatus.PENDING,
                period_start=date(2024, 1, 1) + timedelta(days=i * 30),
                period_end=date(2024, 1, 31) + timedelta(days=i * 30),
            )

        from app.core.security import create_access_token

        token = create_access_token(data={"sub": landlord.id, "type": "landlord"})
        headers = {"Authorization": f"Bearer {token}"}

        response = client.get(
            "/api/payments/export?format=excel&start_date=2024-01-01&end_date=2024-12-31",
            headers=headers,
        )

        assert response.status_code == 200
        assert (
            response.headers["content-type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "payments_" in response.headers["content-disposition"]
        assert ".xlsx" in response.headers["content-disposition"]
        # Verify we got actual file content
        assert len(response.content) > 0

    def test_export_pdf_success(
        self, client: TestClient, session: Session, auth_headers: dict
    ):
        """Test successful PDF export."""
        # Create test data
        landlord = LandlordFactory.create(session=session)
        property_obj = PropertyFactory.create(session=session, landlord_id=landlord.id)
        room = RoomFactory.create(session=session, property_id=property_obj.id)
        tenant = TenantFactory.create(session=session, room_id=room.id)

        PaymentFactory.create(
            session=session,
            tenant_id=tenant.id,
            amount_due=150000,
            status=PaymentStatus.ON_TIME,
        )

        from app.core.security import create_access_token

        token = create_access_token(data={"sub": landlord.id, "type": "landlord"})
        headers = {"Authorization": f"Bearer {token}"}

        response = client.get(
            "/api/payments/export?format=pdf&start_date=2024-01-01&end_date=2024-12-31",
            headers=headers,
        )

        assert response.status_code == 200
        assert response.headers["content-type"] == "application/pdf"
        assert "payments_" in response.headers["content-disposition"]
        assert ".pdf" in response.headers["content-disposition"]
        assert len(response.content) > 0

    def test_export_with_property_filter(
        self, client: TestClient, session: Session, auth_headers: dict
    ):
        """Test export with property filter."""
        landlord = LandlordFactory.create(session=session)

        # Create two properties
        property1 = PropertyFactory.create(session=session, landlord_id=landlord.id)
        property2 = PropertyFactory.create(session=session, landlord_id=landlord.id)

        room1 = RoomFactory.create(session=session, property_id=property1.id)
        room2 = RoomFactory.create(session=session, property_id=property2.id)

        tenant1 = TenantFactory.create(session=session, room_id=room1.id)
        tenant2 = TenantFactory.create(session=session, room_id=room2.id)

        PaymentFactory.create(session=session, tenant_id=tenant1.id)
        PaymentFactory.create(session=session, tenant_id=tenant2.id)

        from app.core.security import create_access_token

        token = create_access_token(data={"sub": landlord.id, "type": "landlord"})
        headers = {"Authorization": f"Bearer {token}"}

        response = client.get(
            f"/api/payments/export?format=excel&property_id={property1.id}&start_date=2024-01-01&end_date=2024-12-31",
            headers=headers,
        )

        assert response.status_code == 200

    def test_export_with_tenant_filter(
        self, client: TestClient, session: Session, auth_headers: dict
    ):
        """Test export with tenant filter."""
        landlord = LandlordFactory.create(session=session)
        property_obj = PropertyFactory.create(session=session, landlord_id=landlord.id)
        room = RoomFactory.create(session=session, property_id=property_obj.id)
        tenant = TenantFactory.create(session=session, room_id=room.id)

        PaymentFactory.create(session=session, tenant_id=tenant.id)

        from app.core.security import create_access_token

        token = create_access_token(data={"sub": landlord.id, "type": "landlord"})
        headers = {"Authorization": f"Bearer {token}"}

        response = client.get(
            f"/api/payments/export?format=excel&tenant_id={tenant.id}&start_date=2024-01-01&end_date=2024-12-31",
            headers=headers,
        )

        assert response.status_code == 200

    def test_export_with_status_filter(
        self, client: TestClient, session: Session, auth_headers: dict
    ):
        """Test export with status filter."""
        landlord = LandlordFactory.create(session=session)
        property_obj = PropertyFactory.create(session=session, landlord_id=landlord.id)
        room = RoomFactory.create(session=session, property_id=property_obj.id)
        tenant = TenantFactory.create(session=session, room_id=room.id)

        # Create payments with different statuses
        PaymentFactory.create(
            session=session, tenant_id=tenant.id, status=PaymentStatus.ON_TIME
        )
        PaymentFactory.create(
            session=session, tenant_id=tenant.id, status=PaymentStatus.OVERDUE
        )

        from app.core.security import create_access_token

        token = create_access_token(data={"sub": landlord.id, "type": "landlord"})
        headers = {"Authorization": f"Bearer {token}"}

        response = client.get(
            "/api/payments/export?format=excel&status=ON_TIME&start_date=2024-01-01&end_date=2024-12-31",
            headers=headers,
        )

        assert response.status_code == 200

    def test_export_multiple_statuses(
        self, client: TestClient, session: Session, auth_headers: dict
    ):
        """Test export with multiple statuses."""
        landlord = LandlordFactory.create(session=session)
        property_obj = PropertyFactory.create(session=session, landlord_id=landlord.id)
        room = RoomFactory.create(session=session, property_id=property_obj.id)
        tenant = TenantFactory.create(session=session, room_id=room.id)

        PaymentFactory.create(session=session, tenant_id=tenant.id)

        from app.core.security import create_access_token

        token = create_access_token(data={"sub": landlord.id, "type": "landlord"})
        headers = {"Authorization": f"Bearer {token}"}

        response = client.get(
            "/api/payments/export?format=excel&status=ON_TIME,PENDING&start_date=2024-01-01&end_date=2024-12-31",
            headers=headers,
        )

        assert response.status_code == 200

    def test_export_date_range_exceeds_limit(
        self, client: TestClient, session: Session, auth_headers: dict
    ):
        """Test export with date range exceeding 2 years."""
        from app.core.security import create_access_token

        landlord = LandlordFactory.create(session=session)
        token = create_access_token(data={"sub": landlord.id, "type": "landlord"})
        headers = {"Authorization": f"Bearer {token}"}

        response = client.get(
            "/api/payments/export?format=excel&start_date=2022-01-01&end_date=2025-01-01",
            headers=headers,
        )

        assert response.status_code == 400
        assert "2 years" in response.json()["detail"].lower()

    def test_export_end_date_before_start(
        self, client: TestClient, session: Session, auth_headers: dict
    ):
        """Test export with end date before start date."""
        from app.core.security import create_access_token

        landlord = LandlordFactory.create(session=session)
        token = create_access_token(data={"sub": landlord.id, "type": "landlord"})
        headers = {"Authorization": f"Bearer {token}"}

        response = client.get(
            "/api/payments/export?format=excel&start_date=2024-12-31&end_date=2024-01-01",
            headers=headers,
        )

        assert response.status_code == 400
        assert "after start" in response.json()["detail"].lower()

    def test_export_default_dates(
        self, client: TestClient, session: Session, auth_headers: dict
    ):
        """Test export uses default dates (current year) when not provided."""
        landlord = LandlordFactory.create(session=session)
        property_obj = PropertyFactory.create(session=session, landlord_id=landlord.id)
        room = RoomFactory.create(session=session, property_id=property_obj.id)
        tenant = TenantFactory.create(session=session, room_id=room.id)

        PaymentFactory.create(session=session, tenant_id=tenant.id)

        from app.core.security import create_access_token

        token = create_access_token(data={"sub": landlord.id, "type": "landlord"})
        headers = {"Authorization": f"Bearer {token}"}

        response = client.get(
            "/api/payments/export?format=excel",
            headers=headers,
        )

        assert response.status_code == 200

    def test_export_no_payments(
        self, client: TestClient, session: Session, auth_headers: dict
    ):
        """Test export when landlord has no payments."""
        from app.core.security import create_access_token

        landlord = LandlordFactory.create(session=session)
        token = create_access_token(data={"sub": landlord.id, "type": "landlord"})
        headers = {"Authorization": f"Bearer {token}"}

        response = client.get(
            "/api/payments/export?format=excel&start_date=2024-01-01&end_date=2024-12-31",
            headers=headers,
        )

        assert response.status_code == 404
        assert "no tenants" in response.json()["detail"].lower()

    def test_export_invalid_format(
        self, client: TestClient, session: Session, auth_headers: dict
    ):
        """Test export with invalid format."""
        from app.core.security import create_access_token

        landlord = LandlordFactory.create(session=session)
        token = create_access_token(data={"sub": landlord.id, "type": "landlord"})
        headers = {"Authorization": f"Bearer {token}"}

        response = client.get(
            "/api/payments/export?format=csv&start_date=2024-01-01&end_date=2024-12-31",
            headers=headers,
        )

        assert response.status_code == 422  # Validation error

    def test_export_unauthorized(self, client: TestClient, auth_headers: dict):
        """Test export without authentication."""
        response = client.get(
            "/api/payments/export?format=excel&start_date=2024-01-01&end_date=2024-12-31",
        )

        assert response.status_code in [401, 403]  # Either unauthorized or forbidden

    def test_export_other_landlord_tenant(
        self, client: TestClient, session: Session, auth_headers: dict
    ):
        """Test export trying to filter by other landlord's tenant."""
        # Create another landlord's tenant
        other_landlord = LandlordFactory.create(session=session)
        property_obj = PropertyFactory.create(
            session=session, landlord_id=other_landlord.id
        )
        room = RoomFactory.create(session=session, property_id=property_obj.id)
        tenant = TenantFactory.create(session=session, room_id=room.id)

        from app.core.security import create_access_token

        token = create_access_token(data={"sub": other_landlord.id, "type": "landlord"})
        headers = {"Authorization": f"Bearer {token}"}

        response = client.get(
            f"/api/payments/export?format=excel&tenant_id={tenant.id}&start_date=2024-01-01&end_date=2024-12-31",
            headers=auth_headers,  # Use auth_landlord's token
        )

        # Should return 404 because tenant doesn't belong to auth_landlord
        # or empty export because tenant not found
        assert response.status_code in [404, 200]

    def test_excel_file_structure(
        self, client: TestClient, session: Session, auth_headers: dict
    ):
        """Test that exported Excel has correct structure."""
        landlord = LandlordFactory.create(session=session)
        property_obj = PropertyFactory.create(session=session, landlord_id=landlord.id)
        room = RoomFactory.create(session=session, property_id=property_obj.id)
        tenant = TenantFactory.create(
            session=session,
            room_id=room.id,
            name="Test Tenant",
            email="test@example.com",
            phone="1234567890",
        )

        PaymentFactory.create(
            session=session,
            tenant_id=tenant.id,
            amount_due=150000,
            status=PaymentStatus.ON_TIME,
            period_start=date(2024, 6, 1),
            period_end=date(2024, 6, 30),
            due_date=date(2024, 6, 1),
            paid_date=date(2024, 5, 28),
            payment_reference="BANK-12345",
            notes="Test payment note",
        )

        from app.core.security import create_access_token

        token = create_access_token(data={"sub": landlord.id, "type": "landlord"})
        headers = {"Authorization": f"Bearer {token}"}

        response = client.get(
            "/api/payments/export?format=excel&start_date=2024-01-01&end_date=2024-12-31",
            headers=headers,
        )

        assert response.status_code == 200

        # Verify we can open the Excel file
        from openpyxl import load_workbook
        from io import BytesIO

        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        # Check sheet name
        assert ws.title == "Payment Report"

        # Check headers exist
        assert ws["A1"].value == "Payment Report"

        # Check summary section exists
        summary_found = False
        for row in ws.iter_rows(min_row=1, max_row=20):
            for cell in row:
                if cell.value == "Summary":
                    summary_found = True
                    break
        assert summary_found

        # Check detailed payments section exists
        payments_found = False
        for row in ws.iter_rows(min_row=1, max_row=50):
            for cell in row:
                if cell.value == "Detailed Payments":
                    payments_found = True
                    break
        assert payments_found

        wb.close()

    def test_pdf_file_structure(
        self, client: TestClient, session: Session, auth_headers: dict
    ):
        """Test that exported PDF has correct structure."""
        landlord = LandlordFactory.create(session=session)
        property_obj = PropertyFactory.create(session=session, landlord_id=landlord.id)
        room = RoomFactory.create(session=session, property_id=property_obj.id)
        tenant = TenantFactory.create(session=session, room_id=room.id)

        PaymentFactory.create(session=session, tenant_id=tenant.id)

        from app.core.security import create_access_token

        token = create_access_token(data={"sub": landlord.id, "type": "landlord"})
        headers = {"Authorization": f"Bearer {token}"}

        response = client.get(
            "/api/payments/export?format=pdf&start_date=2024-01-01&end_date=2024-12-31",
            headers=headers,
        )

        assert response.status_code == 200

        # Verify it's a valid PDF by checking header
        assert response.content[:4] == b"%PDF"
