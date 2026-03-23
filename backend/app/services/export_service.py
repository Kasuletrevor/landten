"""
Export service for generating Excel and PDF reports of payment data.
"""

from collections import defaultdict
from datetime import date
from dataclasses import dataclass
from decimal import Decimal
from io import BytesIO
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import inch
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlmodel import Session
from app.models.payment import Payment, PaymentStatus
from app.models.property import Property
from app.models.room import Room
from app.models.tenant import Tenant


@dataclass(frozen=True)
class PaymentExportRow:
    payment: Payment
    tenant_name: str
    tenant_email: str
    tenant_phone: str
    property_name: str
    room_name: str
    currency: str
    days_overdue: int | None


class ExportService:
    """Service for exporting payment data to various formats."""

    # LandTen brand colors
    PRIMARY_COLOR = "6C5DD3"  # Purple (without # for openpyxl)
    PRIMARY_COLOR_PDF = "#6C5DD3"  # Purple (with # for reportlab)
    SECONDARY_COLOR = "8B5CF6"  # Light purple
    BG_COLOR = "F3F4F6"  # Light gray
    BG_COLOR_PDF = "#F3F4F6"  # Light gray (with # for reportlab)
    TEXT_COLOR = "1F2937"  # Dark gray

    @staticmethod
    def _build_payment_rows(
        session: Session, payments: List[Payment]
    ) -> list[PaymentExportRow]:
        tenant_cache: dict[str, Tenant | None] = {}
        room_cache: dict[str, Room | None] = {}
        property_cache: dict[str, Property | None] = {}
        rows: list[PaymentExportRow] = []

        for payment in payments:
            tenant = tenant_cache.get(payment.tenant_id)
            if payment.tenant_id not in tenant_cache:
                tenant = session.get(Tenant, payment.tenant_id)
                tenant_cache[payment.tenant_id] = tenant

            room: Room | None = None
            if tenant and tenant.room_id:
                room = room_cache.get(tenant.room_id)
                if tenant.room_id not in room_cache:
                    room = session.get(Room, tenant.room_id)
                    room_cache[tenant.room_id] = room

            property_obj: Property | None = None
            if room and room.property_id:
                property_obj = property_cache.get(room.property_id)
                if room.property_id not in property_cache:
                    property_obj = session.get(Property, room.property_id)
                    property_cache[room.property_id] = property_obj

            days_overdue = None
            if payment.status == PaymentStatus.OVERDUE:
                days_overdue = (date.today() - payment.due_date).days

            rows.append(
                PaymentExportRow(
                    payment=payment,
                    tenant_name=tenant.name if tenant else "Unknown",
                    tenant_email=tenant.email if tenant else "",
                    tenant_phone=tenant.phone if tenant else "",
                    property_name=property_obj.name if property_obj else "Unknown",
                    room_name=room.name if room else "Unknown",
                    currency=room.currency if room else "USD",
                    days_overdue=days_overdue,
                )
            )

        return rows

    @staticmethod
    def _sum_amounts_by_currency(
        rows: list[PaymentExportRow],
        statuses: set[PaymentStatus] | None = None,
    ) -> dict[str, Decimal]:
        totals: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
        for row in rows:
            if statuses is not None and row.payment.status not in statuses:
                continue
            totals[row.currency] += Decimal(row.payment.amount_due)
        return dict(totals)

    @staticmethod
    def _build_summary_rows(rows: list[PaymentExportRow]) -> list[list[str | int]]:
        summary_rows: list[list[str | int]] = [["Total Payments", len(rows)]]
        metric_totals = [
            (
                "Total Expected",
                ExportService._sum_amounts_by_currency(rows),
            ),
            (
                "Total Received",
                ExportService._sum_amounts_by_currency(
                    rows, {PaymentStatus.ON_TIME, PaymentStatus.LATE}
                ),
            ),
            (
                "Total Outstanding",
                ExportService._sum_amounts_by_currency(
                    rows, {PaymentStatus.PENDING, PaymentStatus.OVERDUE}
                ),
            ),
            (
                "Total Waived",
                ExportService._sum_amounts_by_currency(rows, {PaymentStatus.WAIVED}),
            ),
        ]

        for label, totals in metric_totals:
            if not totals:
                summary_rows.append([label, ExportService._format_currency(Decimal("0"), "USD")])
                continue

            currencies = sorted(totals)
            if len(currencies) == 1:
                currency = currencies[0]
                summary_rows.append(
                    [label, ExportService._format_currency(totals[currency], currency)]
                )
                continue

            for currency in currencies:
                summary_rows.append(
                    [
                        f"{label} ({currency})",
                        ExportService._format_currency(totals[currency], currency),
                    ]
                )

        return summary_rows

    @staticmethod
    def generate_excel(
        session: Session,
        payments: List[Payment],
        start_date: date,
        end_date: date,
        landlord_name: str,
    ) -> BytesIO:
        """
        Generate Excel workbook with payment data.

        Args:
            payments: List of Payment objects to export
            start_date: Start of date range
            end_date: End of date range
            landlord_name: Name of the landlord for header

        Returns:
            BytesIO: Excel file as bytes buffer
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Payment Report"
        rows = ExportService._build_payment_rows(session, payments)

        # Define styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(
            start_color=ExportService.PRIMARY_COLOR,
            end_color=ExportService.PRIMARY_COLOR,
            fill_type="solid",
        )
        subheader_font = Font(bold=True, size=11)
        subheader_fill = PatternFill(
            start_color=ExportService.BG_COLOR,
            end_color=ExportService.BG_COLOR,
            fill_type="solid",
        )
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Report Header
        ws["A1"] = f"Payment Report"
        ws["A1"].font = Font(bold=True, size=16, color=ExportService.PRIMARY_COLOR)
        ws["A2"] = f"Landlord: {landlord_name}"
        ws["A2"].font = Font(size=11)
        ws["A3"] = (
            f"Period: {start_date.strftime('%B %d, %Y')} - {end_date.strftime('%B %d, %Y')}"
        )
        ws["A3"].font = Font(size=11)
        ws["A4"] = f"Generated: {date.today().strftime('%B %d, %Y')}"
        ws["A4"].font = Font(size=10, italic=True)

        # Empty row
        current_row = 6

        # Summary Section
        ws[f"A{current_row}"] = "Summary"
        ws[f"A{current_row}"].font = Font(
            bold=True, size=14, color=ExportService.PRIMARY_COLOR
        )
        current_row += 1

        summary_data = ExportService._build_summary_rows(rows)

        for label, value in summary_data:
            ws[f"A{current_row}"] = label
            ws[f"A{current_row}"].font = subheader_font
            ws[f"B{current_row}"] = value
            ws[f"B{current_row}"].alignment = Alignment(horizontal="right")
            current_row += 1

        current_row += 2  # Empty rows

        # Detailed Payments Section
        ws[f"A{current_row}"] = "Detailed Payments"
        ws[f"A{current_row}"].font = Font(
            bold=True, size=14, color=ExportService.PRIMARY_COLOR
        )
        current_row += 1

        # Column headers
        headers = [
            "Tenant Name",
            "Email",
            "Phone",
            "Property",
            "Room",
            "Period Start",
            "Period End",
            "Amount Due",
            "Currency",
            "Due Date",
            "Paid Date",
            "Status",
            "Days Overdue",
            "Reference",
            "Notes",
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = border

        current_row += 1
        data_start_row = current_row

        # Data rows
        for row in rows:
            payment = row.payment
            row_data = [
                row.tenant_name,
                row.tenant_email,
                row.tenant_phone,
                row.property_name,
                row.room_name,
                payment.period_start.strftime("%Y-%m-%d"),
                payment.period_end.strftime("%Y-%m-%d"),
                payment.amount_due,
                row.currency,
                payment.due_date.strftime("%Y-%m-%d"),
                payment.paid_date.strftime("%Y-%m-%d") if payment.paid_date else "",
                payment.status.value,
                row.days_overdue if row.days_overdue else "",
                payment.payment_reference or "",
                payment.notes or "",
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = value
                cell.border = border

                if col_num == 8:
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right")
                elif col_num in [12, 13]:
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="left")

            current_row += 1

        # Adjust column widths
        column_widths = [20, 25, 15, 20, 15, 12, 12, 12, 10, 12, 12, 12, 13, 20, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Freeze header row
        ws.freeze_panes = f"A{data_start_row}"

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generate_pdf(
        session: Session,
        payments: List[Payment],
        start_date: date,
        end_date: date,
        landlord_name: str,
    ) -> BytesIO:
        """
        Generate PDF report with payment data.

        Args:
            payments: List of Payment objects to export
            start_date: Start of date range
            end_date: End of date range
            landlord_name: Name of the landlord for header

        Returns:
            BytesIO: PDF file as bytes buffer
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(letter),
            rightMargin=0.5 * inch,
            leftMargin=0.5 * inch,
            topMargin=0.5 * inch,
            bottomMargin=0.5 * inch,
        )
        rows = ExportService._build_payment_rows(session, payments)

        # Container for elements
        elements = []

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=20,
            textColor=colors.HexColor(ExportService.PRIMARY_COLOR_PDF),
            spaceAfter=12,
            alignment=1,  # Center
        )
        subtitle_style = ParagraphStyle(
            "CustomSubtitle",
            parent=styles["Normal"],
            fontSize=11,
            spaceAfter=6,
            alignment=1,  # Center
        )
        summary_style = ParagraphStyle(
            "SummaryTitle",
            parent=styles["Heading2"],
            fontSize=14,
            textColor=colors.HexColor(ExportService.PRIMARY_COLOR_PDF),
            spaceAfter=10,
        )

        # Header
        elements.append(Paragraph("Payment Report", title_style))
        elements.append(Paragraph(f"<b>Landlord:</b> {landlord_name}", subtitle_style))
        elements.append(
            Paragraph(
                f"<b>Period:</b> {start_date.strftime('%B %d, %Y')} - {end_date.strftime('%B %d, %Y')}",
                subtitle_style,
            )
        )
        elements.append(
            Paragraph(
                f"<b>Generated:</b> {date.today().strftime('%B %d, %Y')}",
                subtitle_style,
            )
        )
        elements.append(Spacer(1, 0.2 * inch))

        # Summary Section
        elements.append(Paragraph("Summary", summary_style))

        summary_data = [["Metric", "Value"]]
        for label, value in ExportService._build_summary_rows(rows):
            summary_data.append([str(label), str(value)])

        summary_table = Table(summary_data, colWidths=[2.5 * inch, 2 * inch])
        summary_table.setStyle(
            TableStyle(
                [
                    (
                        "BACKGROUND",
                        (0, 0),
                        (-1, 0),
                        colors.HexColor(ExportService.PRIMARY_COLOR_PDF),
                    ),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 11),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    (
                        "BACKGROUND",
                        (0, 1),
                        (-1, -1),
                        colors.HexColor(ExportService.BG_COLOR_PDF),
                    ),
                    ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                    ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
                ]
            )
        )
        elements.append(summary_table)
        elements.append(Spacer(1, 0.3 * inch))

        # Detailed Payments Section
        if rows:
            elements.append(Paragraph("Detailed Payments", summary_style))

            # Table headers
            table_data = [
                ["Tenant", "Property", "Room", "Period", "Amount", "Due Date", "Status"]
            ]

            for row in rows:
                payment = row.payment
                period = (
                    f"{payment.period_start.strftime('%b %d')} - "
                    f"{payment.period_end.strftime('%b %d')}"
                )
                amount = ExportService._format_currency(payment.amount_due, row.currency)

                table_data.append(
                    [
                        row.tenant_name,
                        row.property_name,
                        row.room_name,
                        period,
                        amount,
                        payment.due_date.strftime("%b %d, %Y"),
                        payment.status.value,
                    ]
                )

            # Create table
            payments_table = Table(
                table_data,
                colWidths=[
                    1.4 * inch,
                    1.2 * inch,
                    0.9 * inch,
                    1.3 * inch,
                    1.1 * inch,
                    1.0 * inch,
                    0.8 * inch,
                ],
            )
            payments_table.setStyle(
                TableStyle(
                    [
                        (
                            "BACKGROUND",
                            (0, 0),
                            (-1, 0),
                            colors.HexColor(ExportService.PRIMARY_COLOR_PDF),
                        ),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("ALIGN", (4, 1), (4, -1), "RIGHT"),
                        ("ALIGN", (6, 1), (6, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 10),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("FONTSIZE", (0, 1), (-1, -1), 9),
                        (
                            "ROWBACKGROUNDS",
                            (0, 1),
                            (-1, -1),
                            [colors.white, colors.HexColor(ExportService.BG_COLOR_PDF)],
                        ),
                    ]
                )
            )
            elements.append(payments_table)
        else:
            elements.append(
                Paragraph(
                    "No payments found for the selected criteria.", styles["Normal"]
                )
            )

        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer

    @staticmethod
    def _format_currency(amount: Decimal | float | None, currency: str) -> str:
        """Format amount with currency code."""
        if amount is None:
            return "-"
        return f"{currency} {amount:,.2f}"
